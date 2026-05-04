#!/usr/bin/env python3
"""
Import FLH member data from Excel into M75 Manager SQLite database.
Merges data from Membres + Médico sheets for best contact info.
"""
import openpyxl
import sqlite3
import re
import os
import sys
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data.db")
XLSX_PATH = os.path.expanduser("~/Desktop/HTML Codes Webseite/GC 2026-04-24-MEMBERSLESCHT 2025-2026.xlsx")

def parse_date(val):
    """Parse various date formats to YYYY-MM-DD."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s in ("None", "///", "xxx", ""):
        return None
    # Try common formats
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d.%m.%Y"]:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None

def clean_text(val):
    """Clean text values."""
    if not val:
        return None
    s = str(val).strip()
    if s in ("None", "///", "xxx", ""):
        return None
    return s[:255] if len(s) > 255 else s

def extract_phone(val):
    """Extract phone number from text."""
    if not val:
        return None
    s = str(val).strip()
    if not s or s == "None":
        return None
    # Remove parenthetical notes
    s = re.sub(r'\([^)]*\)', '', s)
    s = s.strip()
    if not s:
        return None
    return s[:50]

def main():
    print(f"Loading Excel: {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    
    # Load main members sheet
    ws_members = wb[" Membres 2025 -2026"]
    print(f"Members sheet: {ws_members.max_row} rows")
    
    # Load médico sheet for better contact data
    ws_medico = wb["Médico 2026"]
    print(f"Médico sheet: {ws_medico.max_row} rows")
    
    # Build médico lookup by license number
    medico_data = {}
    for row in ws_medico.iter_rows(min_row=2, max_row=ws_medico.max_row, values_only=True):
        license_num = str(row[5]).strip() if row[5] else ""
        if license_num and license_num not in ("None", "xxx", ""):
            medico_data[license_num] = {
                "phone": extract_phone(row[11] or row[12]),
                "email": clean_text(row[13]),
                "birthdate": parse_date(row[9]),
                "medico_date": parse_date(row[8]),
            }
    print(f"Médico entries indexed: {len(medico_data)}")
    
    # Connect to DB
    db = sqlite3.connect(DB_PATH)
    db.execute("PRAGMA journal_mode=WAL")
    db.execute("PRAGMA foreign_keys=OFF")
    
    # Collect unique teams
    teams_map = {}  # name -> id
    existing_teams = {}
    for row in db.execute("SELECT id, name FROM teams"):
        existing_teams[row[1].lower()] = row[0]
    
    # Collect team names from Excel
    excel_teams = set()
    for row in ws_members.iter_rows(min_row=2, max_row=ws_members.max_row, values_only=True):
        team_name = str(row[11]).strip() if row[11] else ""
        cat_name = str(row[9]).strip() if row[9] else ""
        if team_name and team_name not in ("None", "Contact famille", "CAT à compléter", ""):
            excel_teams.add(team_name)
        if cat_name and cat_name not in ("None", "Contact famille", "CAT à compléter", ""):
            excel_teams.add(cat_name)
    
    # Create teams that don't exist
    for team_name in sorted(excel_teams):
        if team_name.lower() not in existing_teams:
            cursor = db.execute(
                "INSERT INTO teams (name, category) VALUES (?, ?)",
                (team_name, "active")
            )
            teams_map[team_name] = cursor.lastrowid
            print(f"  Created team: {team_name} (id={cursor.lastrowid})")
        else:
            teams_map[team_name] = existing_teams[team_name.lower()]
    
    db.commit()
    print(f"Teams: {len(teams_map)} total")
    
    # Import members
    imported = 0
    skipped = 0
    errors = 0
    
    for row in ws_members.iter_rows(min_row=2, max_row=ws_members.max_row, values_only=True):
        try:
            last_name = clean_text(row[0])
            first_name = clean_text(row[1])
            
            if not last_name:
                skipped += 1
                continue
            
            # Skip "Contact famille" entries (parents, not players)
            cat = str(row[9]).strip() if row[9] else ""
            team_name = str(row[11]).strip() if row[11] else ""
            
            if cat in ("Contact famille", "Contact famille M") or team_name == "Contact famille":
                skipped += 1
                continue
            
            # Build member name
            full_name = f"{last_name}, {first_name}" if first_name else last_name
            
            # Get license number
            license_num = str(row[22]).strip() if row[22] else None
            if license_num in ("None", "xxx", "", None):
                license_num = None
            
            # Get birthdate
            birthdate = parse_date(row[32])
            
            # Get contact info
            email = clean_text(row[38])
            phone = extract_phone(row[37] or row[35])
            
            # Merge with médico data if available
            if license_num and license_num in medico_data:
                md = medico_data[license_num]
                if not email and md["email"]:
                    email = md["email"]
                if not phone and md["phone"]:
                    phone = md["phone"]
                if not birthdate and md["birthdate"]:
                    birthdate = md["birthdate"]
            
            # Determine team_id
            team_id = None
            if team_name and team_name in teams_map:
                team_id = teams_map[team_name]
            elif cat and cat in teams_map:
                team_id = teams_map[cat]
            
            # Nationality
            nationality = clean_text(row[2])
            
            # Address
            address = clean_text(row[3])
            postal_code = clean_text(row[4])
            city = clean_text(row[5])
            
            # Position (derive from category)
            position = cat if cat else None
            
            # Check if member already exists by license number
            existing = None
            if license_num:
                existing = db.execute(
                    "SELECT id FROM members WHERE license_number = ?",
                    (license_num,)
                ).fetchone()
            
            if existing:
                # Update existing member
                db.execute("""
                    UPDATE members SET 
                        name = ?, email = ?, phone = ?, birthdate = ?,
                        team_id = ?, license_number = ?
                    WHERE id = ?
                """, (full_name, email, phone, birthdate, team_id, license_num, existing[0]))
            else:
                # Insert new member
                db.execute("""
                    INSERT INTO members (name, email, phone, birthdate, license_number,
                        team_id, address, membership_status)
                    VALUES (?, ?, ?, ?, ?, ?, ?, 'active')
                """, (
                    full_name, email, phone, birthdate, license_num,
                    team_id,
                    f"{address or ''}, {postal_code or ''} {city or ''}".strip(", ")
                ))
            
            imported += 1
            if imported % 1000 == 0:
                db.commit()
                print(f"  Imported {imported} members...")
                
        except Exception as e:
            errors += 1
            if errors <= 10:
                print(f"  Error on row: {e}")
    
    db.commit()
    
    # Print summary
    total_players = db.execute("SELECT COUNT(*) FROM members").fetchone()[0]
    with_license = db.execute("SELECT COUNT(*) FROM members WHERE license_number IS NOT NULL AND license_number != ''").fetchone()[0]
    with_email = db.execute("SELECT COUNT(*) FROM members WHERE email IS NOT NULL AND email != ''").fetchone()[0]
    with_phone = db.execute("SELECT COUNT(*) FROM members WHERE phone IS NOT NULL AND phone != ''").fetchone()[0]
    
    print(f"\n=== IMPORT COMPLETE ===")
    print(f"Imported: {imported}")
    print(f"Skipped (contacts/etc): {skipped}")
    print(f"Errors: {errors}")
    print(f"\nDatabase totals:")
    print(f"  Total members: {total_players}")
    print(f"  With license: {with_license}")
    print(f"  With email: {with_email}")
    print(f"  With phone: {with_phone}")
    
    # Team breakdown
    print(f"\n  Teams:")
    for row in db.execute("""
        SELECT t.name, COUNT(m.id) 
        FROM teams t LEFT JOIN members m ON m.team_id = t.id 
        GROUP BY t.id ORDER BY COUNT(m.id) DESC
    """):
        print(f"    {row[0]}: {row[1]} members")
    
    db.close()
    print("\nDone!")

if __name__ == "__main__":
    main()
