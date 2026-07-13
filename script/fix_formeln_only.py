#!/usr/bin/env python3
"""Minimal Fix: CAT_JOUEUR -> intern 'Tables' + K/N VLOOKUP-Formelen + Datentypen (Zahl)
Tables new-Labelen NICHT verändern (vom User manuell korrigiert)."""
import openpyxl
from pathlib import Path

F = str(Path.home() / 'Desktop' / 'M75_membres_2026_2027_Codes_alt_neu_130726.xlsx')

wb = openpyxl.load_workbook(F)

# ── 1) Datentypen: Tables new A und C als ZAHL ────────────────
tn = wb['Tables new']
na = nc = 0
for r in range(2, tn.max_row + 1):
    a = tn.cell(r, 1).value
    if isinstance(a, str) and a.strip().isdigit():
        tn.cell(r, 1).value = int(a.strip()); na += 1
    c = tn.cell(r, 3).value
    if isinstance(c, str) and c.strip().isdigit():
        tn.cell(r, 3).value = int(c.strip()); nc += 1

# ── 2) CAT_JOUEUR auf intern 'Tables' umbiegen ────────────────
dn = wb.defined_names['CAT_JOUEUR']
dn.value = "'Tables'!$A$1:$B$77"

# ── 3) Membres: K/N Formeln (alter Code -> neuer Code) ─────
ms = wb['Membres 2026_2027']
nk = nn = 0
for r in range(2, ms.max_row + 1):
    l = ms.cell(r, 12).value  # Spalte L
    o = ms.cell(r, 15).value  # Spalte O
    if isinstance(l, str) and l.startswith('='):
        ms.cell(r, 11).value = f"=IFERROR(VLOOKUP(J{r},'Tables new'!$A:$C,3,FALSE),\"\")"  # K
        nk += 1
    if isinstance(o, str) and o.startswith('='):
        ms.cell(r, 14).value = f"=IFERROR(VLOOKUP(M{r},'Tables new'!$A:$C,3,FALSE),\"\")"  # N
        nn += 1

wb.save(F)
print(f"OK. Datentypen konvertiert: A={na}, C={nc}. K-Formeln: {nk}, N-Formeln: {nn}. CAT_JOUEUR -> {dn.value}")
