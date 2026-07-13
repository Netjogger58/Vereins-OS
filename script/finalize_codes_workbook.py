#!/usr/bin/env python3
"""Finalisiert M75_membres_2026_2027_Codes_alt_neu_130726.xlsx (openpyxl, erhält Formatierung):
1) 'Tables new': Labels konsequent HO/FE + neue Nummern (100er) für die Verwaltungs-/Zahlungscodes.
2) Benannten Bereich CAT_JOUEUR -> in-workbook 'Tables'!$A$1:$B$77 umbiegen.
3) 'Membres 2026_2027': K/N mit VLOOKUP-Formeln füllen (alter Code -> neuer Code) wo L/O Formeln haben.
"""
import openpyxl
from pathlib import Path

F = str(Path.home() / 'Desktop' / 'M75_membres_2026_2027_Codes_alt_neu_130726.xlsx')

# Neue Labels (Spalte E) je alter Code – konsequent HO / FE
LABEL = {
    '0': 'noch zu bestimmen (CAT à compléter)',
    '2': 'Seniors HO', '3': 'U21 HO', '4': 'U17 HO', '5': 'U15 HO', '6': 'U13 HO',
    '7': 'U11 HO', '8': 'U9 HO', '9': 'Vétérans HO',
    '10': 'Arbitre HO (21) / FE (41)', '21': 'U7 (Mixte)', '20': 'U4 (Mixte)',
    '1': 'Officiel HO', '11': 'Officiel FE',
    '12': 'Seniors / FE', '14': 'U17 FE', '15': 'U13 FE', '16': 'U15 FE',
    '17': 'U11 FE', '18': 'U9 FE',
    '50': 'Bénévole (allgemein)', '214': 'Bénévole Famille', '215': 'Bénévole avec Licence',
    '102': 'Seniors HO (11) + Arbitre (21)', '109': 'Vétérans HO (20) + Arbitre (21)',
    '112': 'intern gesperrt (global)', '150': 'Comité (HO=1 / FE=3)',
    '188': 'inactif — Lizenz behalten', '200': 'Donateur', '201': 'Donateur licencié',
    '202': 'Membre honoraire',
    '210': 'Contact famille (M)', '211': 'Contact famille (F)', '212': 'Contact famille',
    '213': "Mère / Père d'accueil",
    '220': 'Arrêt temporaire', '221': 'Arrêt temporaire (licencié) — Lizenz behalten',
    '240': 'Ancien membre (ehemalig)', '250': 'Abandon / Abbruch',
    '251': 'Abandon (licencié) — Lizenz behalten', '252': 'Arrêt jeune — Lizenz behalten',
    '253': 'Arrêt jeune + Contact famille löschen', '254': 'inactif', '255': 'Arrêt',
    '256': 'Arrêt — Lizenz gelöscht',
    '300': 'Prêt, sortie Mersch75 (raus)', '301': 'Prêt, entrée Mersch75 (rein)',
    '302': 'Transfert entrant direct', '303': 'Prêt gratuit',
    '304': 'Transfert vers autre club (raus)',
    '976': 'pausiert (Verletzung)', '977': 'blessé — attend rétablissement',
    '980': 'Sponsor — Lizenz behalten', '984': 'ancien, potentiel officiel',
    '978': 'impayé — Zahlung offen; ggf. Officiel/Coach backup',
    '979': 'impayé jeune — Zahlung offen + Lizenz behalten',
    '981': 'impayé — Lizenz gelöscht, in Liste behalten',
    '982': 'impayé — Zahlung offen + Lizenz behalten (potentiel tft)',
    '983': 'Abbruch — Freitext-Kommentar',
    '985': 'Arrêt personne de contact (arrêt jeunes)',
    '991': 'Lizenz gelöscht + aus Memberslist',
    '992': 'impayé — Zahlung offen + Lizenz behalten (remotiver)',
    '993': 'Lizenz gelöscht', '994': 'aus Memberslist entfernen',
    '996': 'jeune inactif, impayé — Zahlung offen + Kommentar',
    '997': 'adulte inactif, impayé — Zahlung offen + Kommentar',
    '999': 'Divers → Freitext-Kommentar',
}
# Neue Nummern (Spalte C) für den Verwaltungs-/Zahlungs-Block (100er)
NEWCODE = {
    '978': 100, '979': 101, '981': 102, '982': 103, '983': 104, '985': 105,
    '991': 106, '992': 107, '993': 108, '994': 109, '995': 110, '996': 111,
    '997': 112, '998': 113, '999': 114,
}
ADMIN_BEREICH = 'Verwaltung / Zahlung / Lizenz (100er)'

wb = openpyxl.load_workbook(F)

# ── 1) Tables new ──────────────────────────────────────────
tn = wb['Tables new']
present = set()
for r in range(2, tn.max_row + 1):
    a = tn.cell(r, 1).value
    if a in (None, ''):
        continue
    code = str(a).strip()
    present.add(code)
    desc = str(tn.cell(r, 2).value or '')
    # Label (Spalte E)
    if code == '19':
        tn.cell(r, 5).value = 'Vétérans FE' if ('éran' in desc.lower() or 'eran' in desc.lower()) else 'U7 FE'
    elif code in LABEL:
        tn.cell(r, 5).value = LABEL[code]
    # Neue Nummer (Spalte C) + Bereich (Spalte D) für Verwaltungscodes
    if code in NEWCODE:
        tn.cell(r, 3).value = NEWCODE[code]
        tn.cell(r, 4).value = ADMIN_BEREICH

# Fehlende Codes 995 / 998 anhängen
append_row = tn.max_row + 1
for code in ('995', '998'):
    if code not in present:
        tn.cell(append_row, 1).value = int(code)
        tn.cell(append_row, 2).value = '(nicht in Original-TABLES — à préciser)'
        tn.cell(append_row, 3).value = NEWCODE[code]
        tn.cell(append_row, 4).value = ADMIN_BEREICH
        tn.cell(append_row, 5).value = '(à préciser)'
        append_row += 1
last_tn = append_row - 1
tn.auto_filter.ref = f'A1:E{last_tn}'

# ── 2) CAT_JOUEUR auf in-workbook 'Tables' umbiegen ────────
dn = wb.defined_names['CAT_JOUEUR']
dn.value = "'Tables'!$A$1:$B$77"

# ── 3) Membres: K/N Formeln (alter Code -> neuer Code) ─────
ms = wb['Membres 2026_2027']
nk = nn = 0
for r in range(2, ms.max_row + 1):
    l = ms.cell(r, 12).value  # Spalte L
    o = ms.cell(r, 15).value  # Spalte O
    if isinstance(l, str) and l.startswith('='):
        ms.cell(r, 11).value = f"=IFERROR(VLOOKUP(J{r},'Tables new'!$A:$C,3,FALSE),\"\")"  # K
        nk += 1
    if isinstance(o, str) and o.startswith('='):
        ms.cell(r, 14).value = f"=IFERROR(VLOOKUP(M{r},'Tables new'!$A:$C,3,FALSE),\"\")"  # N
        nn += 1

wb.save(F)
print(f"OK. Tables new bis Zeile {last_tn}. K-Formeln: {nk}, N-Formeln: {nn}. CAT_JOUEUR -> {dn.value}")
