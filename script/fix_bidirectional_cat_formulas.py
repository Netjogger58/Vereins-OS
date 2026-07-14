#!/usr/bin/env python3
"""Bidirektional CAT-Code-Mapping an M75_membres_2026_2027_Codes_alt_neu_130726.xlsx.

Erstellt ein 'Tables reverse'-Blatt (neuer Code -> alter Code + Bedeutung)
und ein verstecktes 'Originals'-Blatt mit den ursprünglichen AL-Cat-Werten.
In 'Membres 2026_2027' werden J/K/L und M/N/O zu Formeln umgebaut:

- alter Code in J -> neuer Code in K, Bedeutung in L
- neuer Code in K -> alter Code in J, Bedeutung in L
- dito für M/N/O

Damit zirkuläre Bezüge (J bezieht sich auf K, K bezieht sich auf J)
keine Fehler in Excel erzeugen, wird iteratives Rechnen aktiviert.
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

DEFAULT_FILE = Path.home() / "CascadeProjects" / "M75_membres_2026_2027_Codes_alt_neu_130726.xlsx"


def build_reverse_map(tables_new):
    """Aus 'Tables new' ein Mapping neuer Code -> (alter Code, Bedeutung) bauen.
    Bei mehreren alten Codes pro neuem Code gewinnt der letzte Eintrag."""
    reverse = {}
    for r in range(2, tables_new.max_row + 1):
        old = tables_new.cell(r, 1).value
        desc = tables_new.cell(r, 2).value
        new = tables_new.cell(r, 3).value
        label = tables_new.cell(r, 5).value
        if new in (None, ""):
            continue
        # Normalisieren für den Lookup
        old_key = old if old is not None else ""
        # Label: zuerst Erklärung, sonst Original-Beschreibung
        effective_label = label if label not in (None, "") else desc
        reverse[new] = (old_key, effective_label)
    return reverse


def ensure_tables_reverse(wb, reverse_map):
    if "Tables reverse" in wb.sheetnames:
        ws = wb["Tables reverse"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Tables reverse")
    ws.append(["Neuer Code", "Alter Code", "Bedeutung"])
    # Sortieren: numerisch wenn möglich, sonst alphabetisch
    def sort_key(new_code):
        try:
            return (0, int(new_code))
        except (ValueError, TypeError):
            return (1, str(new_code))
    for new_code in sorted(reverse_map.keys(), key=sort_key):
        old, label = reverse_map[new_code]
        ws.append([new_code, old, label])
    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 46
    return ws


def ensure_originals(wb, membres):
    if "Originals" in wb.sheetnames:
        ws = wb["Originals"]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet("Originals")
    ws.append(["Row", "J_original", "M_original"])
    for r in range(2, membres.max_row + 1):
        ws.append([r, None, None])
    ws.sheet_state = "hidden"
    return ws


def main():
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_FILE
    if not path.exists():
        print(f"Datei nicht gefunden: {path}")
        sys.exit(1)

    # 1) Originalwerte auslesen (data_only=True liefert berechnete Werte)
    wb_data = openpyxl.load_workbook(path, data_only=True)
    membres_data = wb_data["Membres 2026_2027"]
    originals = []
    for r in range(2, membres_data.max_row + 1):
        j_val = membres_data.cell(r, 10).value
        m_val = membres_data.cell(r, 13).value
        originals.append((j_val, m_val))
    wb_data.close()

    # 2) Formelarbeitsmappe öffnen
    wb = openpyxl.load_workbook(path)
    membres = wb["Membres 2026_2027"]
    tables_new = wb["Tables new"]

    reverse_map = build_reverse_map(tables_new)
    if not reverse_map:
        print("Keine Reverse-Mapping-Daten in 'Tables new' gefunden.")
        sys.exit(1)

    ensure_tables_reverse(wb, reverse_map)
    orig_sheet = ensure_originals(wb, membres)

    # Originalwerte in das versteckte Blatt schreiben
    for idx, (j_val, m_val) in enumerate(originals, start=2):
        orig_sheet.cell(idx, 2).value = j_val
        orig_sheet.cell(idx, 3).value = m_val

    # Iterative Berechnung aktivieren, damit zirkuläre Bezüge in Excel funktionieren
    wb.calculation.iterate = True
    wb.calculation.iterateCount = 100
    wb.calculation.iterateDelta = 0.001

    # Formeln für J/K/L und M/N/O
    for r in range(2, membres.max_row + 1):
        # J = alter Code (rückwärts aus K, oder Originalwert)
        membres.cell(r, 10).value = (
            f"=IF($K{r}=\"\",Originals!$B{r},IFERROR(VLOOKUP($K{r},'Tables reverse'!$A:$C,2,FALSE),\"\"))"
        )
        # K = neuer Code (vorwärts aus J)
        membres.cell(r, 11).value = (
            f"=IF($J{r}=\"\",\"\",IFERROR(VLOOKUP($J{r},'Tables new'!$A:$C,3,FALSE),\"\"))"
        )
        # L = Bedeutung (vorwärts aus J, oder rückwärts aus K, oder Original-J)
        membres.cell(r, 12).value = (
            f"=IF($J{r}=\"\",IF($K{r}=\"\",IFERROR(VLOOKUP(Originals!$B{r},'Tables new'!$A:$E,5,FALSE),\"\"),IFERROR(VLOOKUP($K{r},'Tables reverse'!$A:$C,3,FALSE),\"\"))"
            f",IFERROR(VLOOKUP($J{r},'Tables new'!$A:$E,5,FALSE),\"\"))"
        )

        # M/N/O – zweiter Block mit identischem Schema
        membres.cell(r, 13).value = (
            f"=IF($N{r}=\"\",Originals!$C{r},IFERROR(VLOOKUP($N{r},'Tables reverse'!$A:$C,2,FALSE),\"\"))"
        )
        membres.cell(r, 14).value = (
            f"=IF($M{r}=\"\",\"\",IFERROR(VLOOKUP($M{r},'Tables new'!$A:$C,3,FALSE),\"\"))"
        )
        membres.cell(r, 15).value = (
            f"=IF($M{r}=\"\",IF($N{r}=\"\",IFERROR(VLOOKUP(Originals!$C{r},'Tables new'!$A:$E,5,FALSE),\"\"),IFERROR(VLOOKUP($N{r},'Tables reverse'!$A:$C,3,FALSE),\"\"))"
            f",IFERROR(VLOOKUP($M{r},'Tables new'!$A:$E,5,FALSE),\"\"))"
        )

    # Schriftgrößen/Formatierung anpassen (optional)
    for col in range(10, 16):
        letter = get_column_letter(col)
        if membres.column_dimensions[letter].width < 12:
            membres.column_dimensions[letter].width = 12

    wb.save(path)
    print(f"Bidirektionale Formeln geschrieben: {path}")
    print(f"Reverse-Mapping-Einträge: {len(reverse_map)}")


if __name__ == "__main__":
    main()
